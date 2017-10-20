import re
from collections import namedtuple
from openpyxl import load_workbook

def create_course_dict():
    """
    Creates a dictionary containing course info.
    Keys: namedtuple of the form ('program, designation')
    Values: namedtuple of the form('name, prereqs, credits')
            prereqs is a tuple of prereqs where each prereq has the same form as the keys
    """
    wb = load_workbook('newcatalog.xlsx')
    catalog = wb.get_sheet_by_name('catalog')
    Course = namedtuple('Course', 'program, designation')
    CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')
    course_dict = {}
    for row in range(1, catalog.max_row + 1):
        key = Course(get_val(catalog, 'A', row), get_val(catalog, 'B', row))
        prereqs = tuple(tuple(get_split_course(prereq) for prereq in prereqs.split())
                   for prereqs in none_split(get_val(catalog, 'E', row)))
        val = CourseInfo(get_val(catalog, 'C', row), tuple(get_val(catalog, 'D', row).split()), prereqs)
        course_dict[key] = val
    return course_dict


def get_split_course(course):
    """
    Parses a course from programdesignation into the ('program, designation') form.
    e.g. 'CS1101' -> ('CS', '1101')
    """
    return tuple(split_course for course_part in re.findall('((?:[A-Z]+-)?[A-Z]+)(.+)', course)
                 for split_course in course_part)


def none_split(val):
    """Handles calling split on a None value by returning the empty list."""
    return val.split(', ') if val else ()


def get_val(catalog, col, row):
    """Returns the value of a cell."""
    return catalog[col + str(row)].value


def print_dict(dict):
    """Simply prints a dictionary's key and values line by line."""
    for key in dict:
        print(key, dict[key])
